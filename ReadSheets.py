import openpyxl

def read_column_from_table(file_path, table_name, column_name):
    wb = openpyxl.load_workbook(file_path) #Workbook
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        if sheet_name == table_name:
            table_data = []
            headers = [cell.value for cell in sheet[1]]
            column_index = headers.index(column_name) + 1
            for row in sheet.iter_rows(min_row=2, min_col=column_index, values_only=True):
                cell_value = row[0]
                if cell_value is not None:
                    table_data.append(cell_value)
            return table_data

def write_to_txt(data, output_file):
    with open(output_file, 'w') as txt_file:
        questions = []
        for item in data: #For when there are multiple questions in a cell separated by newLines '\n'
            while '\n' in item:
                questions.append(item[0:item.find('\n')].strip())
                item = item[item.find('\n')+1:]
            questions.append(item.strip())
        for question in questions:
            stringToWrite = f"{{\"messages\":[{{\"role\":\"system\",\"content\":\"Assume the role of a seasoned expert [CYBERSECURITY EXPERT] with over two decades of experience in CSA STAR. You are recognized for your profound knowledge in [CYBERSECURITY AND DATA PROTECTION]. You are asked to identify the risks and provide recommendations, cybersecurity solutions and cybersecurity products for small businesses based on a negative response to a given question that you are going to be provided. Your response should not only inform but also demonstrate your deep expertise. Your writing should offer unique insights and education on [HiTrust CSF CONTROLS], supplemented by practical, actionable advice on [IMPLEMENTING CYBERSECURITY SOLUTIONS]. Additionally, incorporate [COMPLIANCE WITH GLOBAL DATA PROTECTION REGULATIONS] to provide a comprehensive view and enhance the reader's understanding of the topic's broader context.\"}},{{\"role\":\"assistant\",\"content\":\"{question}\"}},{{\"role\":\"user\",\"content\":\"NO\"}},{{\"role\":\"user\",\"content\":\"Do not provide any introductions just identify risks and provide practical mitigation steps and suggest relevant cybersecurity products or services for compliance with HiTrust CSF. Focus on concise bullet points, and actionable information throughout your response without an introduction or conclusion or wrap-up.\"}}]}}"
            txt_file.write(str(stringToWrite + '\n'))

if __name__ == "__main__":
    excel_file = "C:\\Users\\64mda\\Downloads\\IRM Read Sheets\\CSA_STAR_All_Data.xlsx"  # Update with your Excel file path
    table_name = "Questions"  # Update with the name of the table
    column_name = "Question"  # Update with the name of the column
    output_txt_file = "output.txt"  # Update with the output text file path

    column_data = read_column_from_table(excel_file, table_name, column_name)
    write_to_txt(column_data, output_txt_file)
